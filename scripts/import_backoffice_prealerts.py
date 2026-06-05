#!/usr/bin/env python3
"""Import backoffice prealerts from CSV/XLSX into Supabase.

Usage:
  python scripts/import_backoffice_prealerts.py --file C:\\path\\prealerts.xlsx
  python scripts/import_backoffice_prealerts.py --file C:\\path\\prealerts.csv --truncate
  python scripts/import_backoffice_prealerts.py --file C:\\path\\prealerts.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib import error, request

try:
    import openpyxl
except ImportError:
    openpyxl = None


ROOT_DIR = Path(__file__).resolve().parents[1]
ENV_FILE = ROOT_DIR / ".env.local"


def load_env_file(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip().lstrip("\ufeff")
        value = value.strip().strip('"').strip("'")
        if key and (key not in os.environ or not os.environ.get(key)):
            os.environ[key] = value


def http_json(method: str, url: str, headers: dict[str, str], payload: Any = None) -> Any:
    data_bytes = None
    if payload is not None:
        data_bytes = json.dumps(payload).encode("utf-8")

    req = request.Request(url=url, method=method, headers=headers, data=data_bytes)
    try:
        with request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8")
            return json.loads(body) if body else {}
    except error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {exc.code} calling {url}: {details}") from exc


def normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text.upper()).strip("_")
    return text


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "nan", "nat"}:
        return ""
    return text


def parse_datetime(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.isoformat()

    text = clean_text(value)
    if not text:
        return None

    text = text.replace("/", "-")
    formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%m-%d-%Y",
        "%m-%d-%Y %H:%M",
        "%m-%d-%Y %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            pass

    try:
        return datetime.fromisoformat(text).isoformat()
    except ValueError:
        return None


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def read_xlsx_rows(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for .xlsx imports. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [clean_text(c) for c in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[i] if i < len(headers) else f"COL_{i}": row[i] for i in range(len(row))}
        data_rows.append(item)
    return data_rows


def pick_value(row: dict[str, Any], candidates: list[str]) -> str:
    for candidate in candidates:
        value = row.get(candidate)
        if clean_text(value):
            return clean_text(value)
    return ""


def infer_client(reference_text: str, equipment_text: str, sheet_title: str, customer_text: str = "") -> str:
    text = f"{reference_text} {equipment_text} {sheet_title} {customer_text}".upper()
    if "XIAOMI" in text:
        return "XIAOMI"
    if "CLARO" in text or "OPERADOR" in text or "DISTRIBUIDOR" in text:
        return "CLARO"
    if "RETAIL" in text or "RETEILER" in text:
        return "RETAILER"
    return "UNKNOWN"


def map_row(raw_row: dict[str, Any], row_number: int, default_sheet_title: str) -> dict[str, Any] | None:
    normalized = {normalize_header(k): v for k, v in raw_row.items()}

    reference = pick_value(normalized, [
        "TICKET_DEL_CLIENTE", "TICKET_CLIENTE", "PRE_ALERTA", "PREALERTA", "CORRELATIVO",
        "NO_CASO", "CASO", "TICKET", "REFERENCIA", "FOLIO", "NUMERO_DE_ORDEN",
    ])
    order_number = pick_value(normalized, ["ORDERRY", "NUMERO_ORDEN", "NO_ORDEN", "ORDER", "ORDER_ID", "OT"])
    guide = pick_value(normalized, [
        "NUMEROS_DE_CONDUCE_IMEI_FOLIO", "NUMERO_DE_CONDUCE", "CONDUCE", "GUIA", "NO_GUIA", "TRACKING"
    ])
    imei = pick_value(normalized, ["IMEI", "IMEI_1", "IMEI1", "UID"])
    serial = pick_value(normalized, ["SERIE", "SERIAL", "SN", "NUMERO_DE_SERIE"])
    customer = pick_value(normalized, ["CUSTOMER", "CLIENTE", "CLIENTE_FINAL", "NOMBRE_CLIENTE", "NOMBRE"])
    equipment = pick_value(normalized, ["EQUIPO", "MODELO", "DISPOSITIVO", "DEVICE", "EQUIPMENT_NAME"])
    details = pick_value(normalized, ["DETALLE", "DETALLES", "OBSERVACIONES", "NOTAS", "DESCRIPTION"])
    status = pick_value(normalized, ["ESTADO", "STATUS"])

    request_at = parse_datetime(pick_value(normalized, ["REQUEST_AT", "FECHA_SOLICITUD", "FECHA_INGRESO", "FECHA"]))
    collected_at = parse_datetime(pick_value(normalized, ["COLLECTED_AT", "FECHA_RECOLECCION", "FECHA_COLECTA"]))
    orderry_at = parse_datetime(pick_value(normalized, ["ORDERRY_AT", "FECHA_ORDERRY", "INGRESADO_ORDERRY"]))

    if not any([reference, order_number, guide, imei, serial, customer, equipment]):
        return None

    sheet_title = pick_value(normalized, ["SHEET_TITLE", "HOJA", "TAB", "PAGINA"]) or default_sheet_title
    client = pick_value(normalized, ["CLIENT", "CLIENTE", "CLIENTE_TIPO", "TIPO_CLIENTE"]) or infer_client(reference, equipment, sheet_title, customer)
    client = client.upper() if client else "UNKNOWN"
    if client not in {"CLARO", "XIAOMI", "RETAILER", "UNKNOWN"}:
        client = infer_client(reference, equipment, sheet_title, customer)

    return {
        "client": client,
        "sheet_title": sheet_title,
        "row_number": row_number,
        "customer": customer,
        "reference": reference,
        "order_number": order_number,
        "guide": guide,
        "imei": imei,
        "serial": serial,
        "equipment_name": equipment,
        "details": details,
        "request_at": request_at,
        "collected_at": collected_at,
        "orderry_at": orderry_at,
        "status": status,
        "raw": {k: clean_text(v) for k, v in raw_row.items()},
    }


def chunked(values: list[dict[str, Any]], size: int) -> list[list[dict[str, Any]]]:
    return [values[i : i + size] for i in range(0, len(values), size)]


def truncate_table(supabase_url: str, service_key: str) -> None:
    url = f"{supabase_url}/rest/v1/backoffice_prealerts?id=not.is.null"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    http_json("DELETE", url, headers)


def insert_rows(supabase_url: str, service_key: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0

    url = f"{supabase_url}/rest/v1/backoffice_prealerts"
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    written = 0
    for batch in chunked(rows, 300):
        result = http_json("POST", url, headers, payload=batch)
        written += len(result) if isinstance(result, list) else len(batch)
    return written


def main() -> int:
    parser = argparse.ArgumentParser(description="Import backoffice prealerts into Supabase")
    parser.add_argument("--file", required=True, help="Path to CSV or XLSX file")
    parser.add_argument("--sheet", default=None, help="Sheet name for XLSX (optional)")
    parser.add_argument("--sheet-title", default="Supabase", help="Default sheet_title value")
    parser.add_argument("--truncate", action="store_true", help="Delete existing rows before import")
    parser.add_argument("--dry-run", action="store_true", help="Parse and map rows without writing")
    args = parser.parse_args()

    load_env_file(ENV_FILE)

    supabase_url = os.getenv("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    if not supabase_url or not service_key:
        raise RuntimeError("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    source_path = Path(args.file)
    if not source_path.exists():
        raise FileNotFoundError(f"Input file not found: {source_path}")

    if source_path.suffix.lower() == ".csv":
        raw_rows = read_csv_rows(source_path)
    elif source_path.suffix.lower() in {".xlsx", ".xlsm"}:
        raw_rows = read_xlsx_rows(source_path, args.sheet)
    else:
        raise RuntimeError("Unsupported file format. Use CSV or XLSX.")

    mapped_rows: list[dict[str, Any]] = []
    for idx, raw in enumerate(raw_rows, start=2):
        mapped = map_row(raw, idx, args.sheet_title)
        if mapped:
            mapped_rows.append(mapped)

    print(f"Input rows: {len(raw_rows)}")
    print(f"Mapped rows: {len(mapped_rows)}")

    if mapped_rows:
        preview = mapped_rows[0]
        print("First mapped row preview:")
        print(json.dumps(preview, ensure_ascii=True, indent=2))

    if args.dry_run:
        print("Dry-run enabled, no rows written.")
        return 0

    if args.truncate:
        truncate_table(supabase_url, service_key)
        print("Existing rows deleted from backoffice_prealerts.")

    written = insert_rows(supabase_url, service_key, mapped_rows)
    print(f"Import completed. Rows written: {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
